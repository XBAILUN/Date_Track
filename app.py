#!/usr/bin/env python
# coding: utf-8

# In[ ]:


from flask import Flask, request, send_file, render_template, flash
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import tempfile
import os

app = Flask(__name__)
app.secret_key = 'your_very_secret_key'

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        file = request.files['file']
        column_name = request.form['column_name']
        data_start_row = int(request.form['data_start_row'])
        header_row = int(request.form['header_row'])
        
        if file:
            temp_dir = tempfile.mkdtemp()
            temp_path = os.path.join(temp_dir, file.filename)
            file.save(temp_path)
            
            # 生成一个临时文件对象，而不是直接生成文件路径
            temp_file = process_excel_file(temp_path, column_name, data_start_row, header_row)
            file_path = temp_file.name  # 获取临时文件的路径
            temp_file.close()  # 关闭临时文件对象，但不删除文件

            response = send_file(file_path, as_attachment=True, download_name='processed_file.xlsx')
            
            # 文件发送后，执行删除操作
            try:
                os.remove(file_path)  # 删除临时文件
            except Exception as e:
                app.logger.error(f"Error removing temporary file: {e}")
            
            return response
        else:
            flash('No file part')
            return redirect(request.url)

    return render_template("index.html")

def process_excel_file(filepath, column_name, data_start_row, header_row):
    wb = openpyxl.load_workbook(filepath)
    today = datetime.today()

    summary_sheet_names = ['Upcoming One Month', 'One Month to Two Months', 'Two Months to Three Months']
    data_to_add = {name: {} for name in summary_sheet_names}
    data_row_counters = {name: 0 for name in summary_sheet_names}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        milestone_column = None

        for cell in sheet[header_row]:
            if cell.value == column_name:
                milestone_column = cell.column
                break

        if not milestone_column:
            continue

        for row in sheet.iter_rows(min_row=data_start_row):
            cell = row[milestone_column - 1]
            if cell.value and isinstance(cell.value, datetime):
                days_diff = (cell.value - today).days
                summary_sheet_name, fill_color = check_date_range_and_get_info(days_diff)
                if summary_sheet_name:
                    cell.fill = fill_color
                    data_to_add[summary_sheet_name].setdefault(sheet_name, []).append([cell.value for cell in row])
                    data_row_counters[summary_sheet_name] += 1

    for sheet_name in summary_sheet_names:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        wb.create_sheet(sheet_name)
        summary_sheet = wb[sheet_name]

        for src_sheet_name, rows in data_to_add[sheet_name].items():
            summary_sheet.append(['Data from: ' + src_sheet_name])
            col_names = [cell.value for cell in wb[src_sheet_name][header_row]]
            summary_sheet.append(col_names)
            for row in rows:
                summary_sheet.append(row)

        summary_sheet.insert_rows(1)
        summary_sheet['A1'] = f'Total Data Rows: {data_row_counters[sheet_name]}'

    # Create a NamedTemporaryFile and return it instead of the path
    temp_file = tempfile.NamedTemporaryFile(mode='w+b', suffix='.xlsx', delete=False)
    wb.save(temp_file.name)
    wb.close()

    return temp_file

def check_date_range_and_get_info(days_diff):
    if days_diff <= 30:
        return 'Upcoming One Month', PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    elif 30 < days_diff <= 60:
        return 'One Month to Two Months', PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    elif 60 < days_diff <= 90:
        return 'Two Months to Three Months', PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    return None, None

if __name__ == "__main__":
    app.run()


# In[ ]:





# In[ ]:




